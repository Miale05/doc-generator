import os
import shutil
from datetime import datetime
import openpyxl
import subprocess
import platform
import argparse

def main():
    # Parsear argumentos de línea de comandos
    parser = argparse.ArgumentParser(description='Generador de Actas desde Excel')
    parser.add_argument('--fila-inicio', type=int, default=2,
                        help='Fila de inicio (default: 2)')
    parser.add_argument('--fila-fin', type=int, default=None,
                        help='Fila de fin (default: última fila)')
    
    args = parser.parse_args()
    
    resumen_path = 'input/Datos.xlsx'
    fila_inicio = args.fila_inicio
    fila_fin = args.fila_fin
    
    # Obtener la fecha de hoy
    today = datetime.now()
    date_str = today.strftime('%Y%m%d')  # Formato YYYYMMDD

    # Crear la carpeta en resultados con la fecha
    result_dir = os.path.join('resultados', date_str)
    os.makedirs(result_dir, exist_ok=True)

    # Leer el archivo Datos.xlsx
    wb_resumen = openpyxl.load_workbook(resumen_path)
    sheet_resumen = wb_resumen.active
    
    # Si no se especifica fila_fin, usar la última fila con datos
    if fila_fin is None:
        fila_fin = sheet_resumen.max_row

    # Procesar filas en el rango especificado
    for row_idx, row in enumerate(sheet_resumen.iter_rows(min_row=fila_inicio, max_row=fila_fin, values_only=True), start=fila_inicio):
        # Asumir columnas en el orden dado (ajustar según el formato real del Excel)
        id_cliente = row[0]
        hora_de_inicio = row[1]
        hora_de_finalizacion = row[2]
        correo_electronico = row[3]
        nombre = row[4]
        nombre_cliente = row[5]
        dni_cliente = row[6]
        fecha_reunion = row[7]
        nombre_de_asesores_externos = row[8]
        cantidad_horas_usadas = row[9]
        tema_1_tratado = row[10]
        tema_2_tratado = row[11]
        tema_3_tratado = row[12]
        tema_otros_tratado = row[13]
        acuerdo_1 = row[14]
        acuerdo_2 = row[15]
        acuerdo_3 = row[16]
        acuerdo_4 = row[17]
        tema_pendiente_1 = row[18]
        tema_pendiente_2 = row[19]
        tema_pendiente_3 = row[20]
        tema_pendiente_4 = row[21]
        tener_otra_reunion = row[22]
        fecha_nueva_reunion = row[23]
        motivo_reunion = row[24]
        nombre_asesor_sura = row[25]
        correo_electronico_asesor_sura = row[26]

        # Generar timestamp único para cada archivo (segundos y milisegundos)
        timestamp = date_str + "_" + datetime.now().strftime('%H%M%S%f')  # HHMMSSffffff

        # Nombre del archivo: <ID>_Acta_<timestamp>.xlsx
        acta_name = f"{id_cliente}_Acta_Reunion_{timestamp}.xlsx"
        acta_path = os.path.join(result_dir, acta_name)

        # Copiar el modelo
        modelo_path = 'modelo/Plantilla.xlsx'
        shutil.copy(modelo_path, acta_path)

        # Editar el archivo copiado
        wb_acta = openpyxl.load_workbook(acta_path)
        sheet_acta = wb_acta.active

        # Configurar el tamaño de página a A2 (Código 66)
        sheet_acta.page_setup.paperSize = 8

        # Solución al error AttributeError: configuramos el ajuste directamente en la hoja
        sheet_acta.sheet_properties.pageSetUpPr.fitToPage = True
        sheet_acta.page_setup.fitToWidth = 1
        sheet_acta.page_setup.fitToHeight = 0 # 0 permite que el alto crezca según el contenido

        # Rellenar campos
        sheet_acta['E8'] = str(nombre_cliente)
        sheet_acta['E10'] = str(dni_cliente)
        # Formatear fecha para mostrar solo día/mes/año
        if isinstance(fecha_reunion, datetime):
            sheet_acta['E12'] = fecha_reunion.strftime('%d/%m/%Y')
        else:
            sheet_acta['E12'] = str(fecha_reunion)
            
        sheet_acta['E14'] = str(nombre_de_asesores_externos)

        if cantidad_horas_usadas == '1 hora':
            sheet_acta['B20'] = 'X'
        if cantidad_horas_usadas == '1.5 horas':
            sheet_acta['B22'] = 'X'
        if cantidad_horas_usadas == '2 horas':
            sheet_acta['B24'] = 'X'
        if cantidad_horas_usadas == '2.5 horas':
            sheet_acta['B26'] = 'X'
        if cantidad_horas_usadas == '3 horas':
            sheet_acta['B28'] = 'X'

        sheet_acta['E31'] = str(tema_1_tratado)
        sheet_acta['E32'] = str(tema_2_tratado)
        sheet_acta['E33'] = str(tema_3_tratado)
        sheet_acta['E34'] = str(tema_otros_tratado)

        sheet_acta['E37'] = str(acuerdo_1)
        sheet_acta['E38'] = str(acuerdo_2)
        sheet_acta['E39'] = str(acuerdo_3)
        sheet_acta['E40'] = str(acuerdo_4)

        sheet_acta['E43'] = str(tema_pendiente_1)
        sheet_acta['E44'] = str(tema_pendiente_2)
        sheet_acta['E45'] = str(tema_pendiente_3)
        sheet_acta['E46'] = str(tema_pendiente_4)

        if tener_otra_reunion == 'Sí':
            sheet_acta['B51'] = 'X'
            # Formatear fecha de nueva reunión si existe
            if isinstance(fecha_nueva_reunion, datetime):
                sheet_acta['H55'] = fecha_nueva_reunion.strftime('%d/%m/%Y')
            else:
                sheet_acta['H55'] = str(fecha_nueva_reunion)
            sheet_acta['F57'] = str(motivo_reunion)
        else:
            sheet_acta['B53'] = 'X'

        sheet_acta['E62'] = str(nombre_asesor_sura)
        sheet_acta['E63'] = str(correo_electronico_asesor_sura)

        # Guardar el archivo Excel
        wb_acta.save(acta_path)

        # Convertir a PDF usando LibreOffice
        pdf_name = acta_name.replace('.xlsx', '.pdf')
        pdf_path = os.path.join(result_dir, pdf_name)
        
        try:
            # Detectar ubicación de LibreOffice según el SO
            if platform.system() == "Windows":
                libreoffice_path = r"C:\Program Files\LibreOffice\program\soffice.exe"
            elif platform.system() == "Darwin":  # macOS
                libreoffice_path = "/Applications/LibreOffice.app/Contents/MacOS/soffice"
            else:  # Linux
                libreoffice_path = "soffice"
            
            # Convertir Excel a PDF
            subprocess.run(
                [libreoffice_path, '--headless', '--convert-to', 'pdf', '--outdir', result_dir, acta_path],
                check=True,
                capture_output=True
            )
        except (subprocess.CalledProcessError, FileNotFoundError) as e:
            print(f"Advertencia: No se pudo convertir {acta_name} a PDF. Asegúrate de que LibreOffice esté instalado.")

    print("Proceso completado.")

if __name__ == "__main__":
    main()